import openpyxl

from backend.core.period_utils import (
    normalize,
    parse_period,
    first_day_of_month,
    subtract_months,
)

from backend.core.excel_utils import (
    find_first_header_row,
    get_headers,
    get_value,
    find_period_row,
    find_last_period_row,
    flag_inconsistency,
    keep_only_relevant_sheets,
    is_empty_period_row,
)

from backend.processors.calculation_processor import (
    resolve_cell_value,
    values_different,
    numeric_difference,
)

from backend.core.excel_format_utils import (
    copy_cell_format,
    copy_column_format_from_source,
    copy_placeholder_formats_from_target,

)

from backend.core.excel_layout_utils import (
    find_previous_source_header_in_target,
)


def format_new_period_row(target_ws, target_row, period_col, period_date):
    template_row = target_row - 1

    if template_row >= 1:
        for col in range(1, target_ws.max_column + 1):
            copy_cell_format(
                target_ws.cell(row=template_row, column=col),
                target_ws.cell(row=target_row, column=col),
            )

    period_cell = target_ws.cell(row=target_row, column=period_col)
    period_cell.value = first_day_of_month(period_date)
    period_cell.number_format = "mmm-yy"


def add_new_source_columns_to_output(
    source_ws,
    target_ws,
    source_header_row,
    target_header_row,
    source_headers,
    target_headers,
    start_date,
    end_date,
    config,
    source_last_period_row,
):
    
    row_offset = target_header_row - source_header_row

    first_placeholder_row = (
        source_last_period_row
        + row_offset
        + 1
    )

    new_columns = []

    mapped_source_columns = {
        normalize(source_col)
        for source_col in config["column_map"].values()
    }
    mapped_source_columns.add(normalize("Period"))

    for cell in source_ws[source_header_row]:
        if not cell.value:
            continue

        source_header_name = str(cell.value).strip()
        normalized_source_name = normalize(source_header_name)

        if normalized_source_name in mapped_source_columns:
            continue

        source_col_num = cell.column

        if normalized_source_name not in target_headers:
            previous_source_col, previous_target_col = find_previous_source_header_in_target(
                source_ws,
                source_header_row,
                source_col_num,
                target_headers,
            )

            if previous_target_col is None:
                raise ValueError(
                    f"Could not find where to insert new column: {source_header_name}"
                )

            new_col_num = previous_target_col + 1

            for source_gap_col in range(previous_source_col + 1, source_col_num):
                source_gap_value = source_ws.cell(
                    row=source_header_row,
                    column=source_gap_col,
                ).value

                if source_gap_value not in ("", None):
                    continue

                if target_ws.cell(row=target_header_row, column=new_col_num).value not in ("", None):
                    target_ws.insert_cols(new_col_num)

                copy_column_format_from_source(
                    source_ws,
                    target_ws,
                    source_gap_col,
                    new_col_num,
                    source_header_row,
                    target_header_row,
                    source_last_row=source_last_period_row,
                )

                new_col_num += 1

            target_ws.insert_cols(new_col_num)

            copy_column_format_from_source(
                source_ws,
                target_ws,
                source_col_num,
                new_col_num,
                source_header_row,
                target_header_row,
                source_last_row=source_last_period_row,
            )

            period_col = target_headers.get(normalize("Period"))

            if not period_col:
                raise ValueError("Could not find Period column in target file")
            
            copy_placeholder_formats_from_target(
                target_ws=target_ws,
                template_col=previous_target_col,
                target_col=new_col_num,
                period_col=period_col,
                first_placeholder_row=first_placeholder_row,
            )

            target_ws.cell(
                row=target_header_row,
                column=new_col_num,
            ).value = source_header_name

            source_separator_col = source_col_num + 1

            if (
                source_separator_col <= source_ws.max_column
                and source_ws.cell(row=source_header_row, column=source_separator_col).value in ("", None)
            ):
                target_separator_col = new_col_num + 1

                if target_ws.cell(row=target_header_row, column=target_separator_col).value not in ("", None):
                    target_ws.insert_cols(target_separator_col)

                copy_column_format_from_source(
                    source_ws,
                    target_ws,
                    source_separator_col,
                    target_separator_col,
                    source_header_row,
                    target_header_row,
                    source_last_row=source_last_period_row,
                )

            target_headers.clear()
            target_headers.update(get_headers(target_ws, target_header_row))

            config["column_map"][source_header_name] = source_header_name
            new_columns.append(source_header_name)

        else:
            new_col_num = target_headers[normalized_source_name]
            config["column_map"][source_header_name] = source_header_name

        period_col = target_headers.get(normalize("Period"))

        if not period_col:
            raise ValueError("Could not find Period column after inserting new column")

        for row in range(target_header_row + 1, target_ws.max_row + 1):
            period_value = target_ws.cell(row=row, column=period_col).value
            period_date = parse_period(period_value)

            if period_date is None:
                continue

            period_month = first_day_of_month(period_date)

            if period_month <= end_date:
                if target_ws.cell(row=row, column=new_col_num).value in ("", None):
                    target_ws.cell(row=row, column=new_col_num).value = 0

    return new_columns


def build_messages(
    config,
    start_date,
    end_date,
    new_columns,
    skipped_columns,
    inconsistencies,
    saved,
):
    messages = []

    if saved:
        messages.append(f"Success. Saved as {config['target_output_file']}")
    else:
        messages.append("Preview generated. No files were saved yet.")

    messages.append(
        f"This application compared and transferred data only within the "
        f"last {config['lookback_months']} months "
        f"({start_date.strftime('%b %y')} to {end_date.strftime('%b %y')})."
    )

    if new_columns:
        messages.append("New source columns detected and added to output:")
        for col in new_columns:
            messages.append(f"- {col}")
    else:
        messages.append("No new columns were detected.")

    if skipped_columns:
        messages.append("Missing columns:")
        for col in sorted(skipped_columns):
            messages.append(f"- {col}")
    else:
        messages.append("No columns were skipped.")

    if inconsistencies:
        messages.append("Cells with inconsistencies were NOT overwritten.")
        messages.append("Inconsistencies found:")

        for item in inconsistencies:
            messages.append(
                f"- {item['period']} | {item['column']} "
                f"target {item['target_cell']} / source {item['source_cell']}: "
                f"target had {item['existing_target_value']}, "
                f"source has {item['new_source_value']}"
            )
    else:
        messages.append("No inconsistencies found.")

    return messages

def find_original_placeholder_rows(
    target_ws,
    target_header_row,
    target_headers,
    column_names,
):
    period_col = target_headers.get(normalize("Period"))

    if not period_col:
        raise ValueError("Could not find Period column in target file")

    placeholder_periods = set()

    for row in range(target_header_row + 1, target_ws.max_row + 1):
        period_value = target_ws.cell(
            row=row,
            column=period_col,
        ).value

        period_date = parse_period(period_value)

        if period_date is None:
            continue

        if is_empty_period_row(
            target_ws,
            row,
            target_headers,
            column_names,
        ):
            placeholder_periods.add(first_day_of_month(period_date))

    return placeholder_periods

def run_electricity_transfer(config, save_outputs=False):
    skipped_columns = set()
    inconsistencies = []
    updated_rows_preview = []

    config = dict(config)
    config["column_map"] = dict(config["column_map"])

    source_wb = openpyxl.load_workbook(config["source_file"], data_only=True)
    source_wb_highlight = openpyxl.load_workbook(config["source_file"], data_only=True)
    target_wb = openpyxl.load_workbook(config["target_file"])

    source_ws = source_wb[config["source_sheet"]]
    source_ws_highlight = source_wb_highlight[config["source_sheet"]]
    target_ws = target_wb[config["target_sheet"]]

    source_header_row = find_first_header_row(source_ws, "Period")
    target_header_row = find_first_header_row(target_ws, "Period")

    source_headers = get_headers(source_ws, source_header_row)
    target_headers = get_headers(target_ws, target_header_row)

    source_period_col = source_headers.get(normalize("Period"))

    if not source_period_col:
        raise ValueError("Could not find Period column in source file")
    
    last_period_row = find_last_period_row(
        source_ws,
        source_period_col,
        source_header_row,
    )

    source_months = []

    for source_row in range(source_header_row + 1, last_period_row + 1):
        period_value = source_ws.cell(row=source_row, column=source_period_col).value
        period_date = parse_period(period_value)

        if period_date is not None:
            source_months.append(first_day_of_month(period_date))

    if not source_months:
        raise ValueError("No valid periods found in source file")

    end_date = max(source_months)
    start_date = subtract_months(end_date, config["lookback_months"])

    source_headers = get_headers(source_ws, source_header_row)
    target_headers = get_headers(target_ws, target_header_row)

    original_placeholder_periods = find_original_placeholder_rows(
        target_ws=target_ws,
        target_header_row=target_header_row,
        target_headers=target_headers,
        column_names=config["column_map"].keys(),
    )

    new_columns = add_new_source_columns_to_output(
        source_ws,
        target_ws,
        source_header_row,
        target_header_row,
        source_headers,
        target_headers,
        start_date,
        end_date,
        config,
        last_period_row,
    )

    period_col = target_headers.get(normalize("Period"))

    if not period_col:
        raise ValueError("Could not find Period column in target file")

    for source_row in range(source_header_row + 1, last_period_row + 1):
        period_value = source_ws.cell(row=source_row, column=source_period_col).value
        period_date = parse_period(period_value)

        if period_date is None:
            continue

        period_month = first_day_of_month(period_date)

        if not (start_date <= period_month <= end_date):
            continue

        target_row = find_period_row(
            target_ws,
            target_headers,
            period_value,
        )

        row_was_added = target_row == target_ws.max_row + 1

        if row_was_added:
            row_was_placeholder = False

            format_new_period_row(
                target_ws,
                target_row,
                period_col,
                period_date,
            )
        else:
            row_was_placeholder = (
                period_month in original_placeholder_periods
            )

            period_cell = target_ws.cell(
                row=target_row,
                column=period_col,
            )
            period_cell.value = first_day_of_month(period_date)
            period_cell.number_format = "mmm-yy"

        is_new_row = row_was_added or row_was_placeholder

        row_preview = {
            "period": period_value,
            "target_row": target_row,
            "is_new_row": is_new_row,
            "row_status": (
                "added"
                if row_was_added
                else "placeholder_filled"
                if row_was_placeholder
                else "existing"
            ),
            "values": {},
        }

        for target_col, source_col in config["column_map"].items():
            source_value = get_value(
                source_ws,
                source_row,
                source_headers,
                source_col,
                skipped_columns,
            )

            target_col_num = target_headers.get(normalize(target_col))

            if not target_col_num:
                skipped_columns.add(target_col)
                continue

            target_cell = target_ws.cell(row=target_row, column=target_col_num)
            old_value = resolve_cell_value(target_ws, target_cell.value)

            if (
                target_col not in new_columns
                and old_value not in ("", None)
                and source_value not in ("", None)
                and values_different(old_value, source_value)
            ):
                diff = numeric_difference(old_value, source_value)

                if diff is None or diff >= 1:
                    flag_inconsistency(
                        target_ws,
                        source_ws_highlight,
                        target_row,
                        target_col_num,
                        source_row,
                        source_headers.get(normalize(source_col)),
                        inconsistencies,
                        period_value,
                        target_col,
                        old_value,
                        source_value,
                    )

                    target_cell.value = source_value

                    if values_different(old_value, source_value):
                        row_preview["values"][target_col] = {
                            "cell": target_cell.coordinate,
                            "status": "updated",
                            "old_value": old_value,
                            "new_value": source_value,
                            "final_value": source_value,
                        }

                    continue

            if source_value in ("", None):
                row_preview["values"][target_col] = {
                    "cell": target_cell.coordinate,
                    "status": "blank_source_skipped",
                    "old_value": old_value,
                    "new_value": source_value,
                    "final_value": old_value,
                }
            else:
                target_cell.value = source_value

                diff = numeric_difference(old_value, source_value)
                changed = values_different(old_value, source_value)

                if is_new_row:
                    status = "new_row_value"
                elif changed and diff is not None and diff < 1:
                    status = "rounding_update"
                elif changed:
                    status = "updated"
                else:
                    status = "unchanged"

                row_preview["values"][target_col] = {
                    "cell": target_cell.coordinate,
                    "status": status,
                    "old_value": old_value,
                    "new_value": source_value,
                    "final_value": source_value,
                }

        if row_preview["values"]:
            updated_rows_preview.append(row_preview)

    keep_only_relevant_sheets(target_wb, config["target_sheet"])

    if save_outputs:
        target_wb.save(config["target_output_file"])

    messages = build_messages(
        config=config,
        start_date=start_date,
        end_date=end_date,
        new_columns=new_columns,
        skipped_columns=skipped_columns,
        inconsistencies=inconsistencies,
        saved=save_outputs,
    )

    return {
        "updated_rows": updated_rows_preview,
        "new_columns": new_columns,
        "skipped_columns": sorted(skipped_columns),
        "inconsistencies": inconsistencies,
        "messages": messages,
        "period_range": {
            "start": start_date.strftime("%Y-%m-%d"),
            "end": end_date.strftime("%Y-%m-%d"),
            "label": f"{start_date.strftime('%b %y')} to {end_date.strftime('%b %y')}",
        },
        "saved": save_outputs,
    }


def preview_electricity_transfer(config):
    return run_electricity_transfer(config, save_outputs=False)


def apply_electricity_transfer(config):
    return run_electricity_transfer(
        config,
        save_outputs=True,
    )

def process_electricity(config, preview_mode=False):
    result = run_electricity_transfer(config, save_outputs=not preview_mode)

    for message in result["messages"]:
        print(message)

    return result