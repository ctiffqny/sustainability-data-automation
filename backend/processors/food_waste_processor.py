from __future__ import annotations

from collections import defaultdict
from copy import copy
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
import shutil
from typing import Any

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from backend.core.pdf_config_utils import PDFConfigLoader
from backend.processors.pdf_processor import PDFProcessor
from backend.processors.calculation_processor import data_processing


PREVIEW_FILL = PatternFill(
    fill_type="solid",
    fgColor="FFF2CC",
)


class FoodWasteTransferError(Exception):
    """Raised when food-waste data cannot be transferred."""


def _normalize_header(value: Any) -> str:
    return " ".join(str(value or "").strip().casefold().split())


def _normalize_period(value: Any) -> str | None:
    if value is None or value == "":
        return None

    if isinstance(value, (datetime, date)):
        return value.strftime("%b-%y")

    text = str(value).strip()

    try:
        return PDFProcessor.normalize_month(text)
    except ValueError:
        return None


def _decimal_to_number(value: Decimal | None) -> float | None:
    return float(value) if value is not None else None


def _copy_workbook(source: str | Path, output: str | Path) -> Path:
    source_path = Path(source)
    output_path = Path(output)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_path, output_path)

    return output_path


def _find_header_columns(
    worksheet,
    header_row: int,
) -> dict[str, int]:
    columns: dict[str, int] = {}

    for cell in worksheet[header_row]:
        normalized = _normalize_header(cell.value)

        if normalized:
            columns[normalized] = cell.column

    return columns


def _find_period_row(
    worksheet,
    *,
    period_column: int,
    data_start_row: int,
    month: str,
) -> int | None:
    for row_number in range(
        data_start_row,
        worksheet.max_row + 1,
    ):
        value = worksheet.cell(
            row=row_number,
            column=period_column,
        ).value

        if _normalize_period(value) == month:
            return row_number

    return None


def _next_period_row(
    worksheet,
    *,
    period_column: int,
    data_start_row: int,
) -> int:
    row_number = max(data_start_row, worksheet.max_row + 1)

    while worksheet.cell(
        row=row_number,
        column=period_column,
    ).value not in (None, ""):
        row_number += 1

    return row_number


def _copy_row_format(
    worksheet,
    source_row: int,
    target_row: int,
) -> None:
    for column in range(1, worksheet.max_column + 1):
        source_cell = worksheet.cell(
            row=source_row,
            column=column,
        )
        target_cell = worksheet.cell(
            row=target_row,
            column=column,
        )

        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)

        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format

        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)

        if source_cell.border:
            target_cell.border = copy(source_cell.border)

        if source_cell.font:
            target_cell.font = copy(source_cell.font)

        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)


def _extract_pdf_values(
    config: dict[str, Any],
) -> tuple[str, dict[str, Decimal], list[dict[str, Any]]]:
    source_files = config.get("source_files") or []
    month = PDFProcessor.normalize_month(config["month"])

    if not source_files:
        raise FoodWasteTransferError(
            "No food-waste PDF source files were supplied."
        )

    pdf_processor = PDFProcessor(
        config_path=config.get("config_path")
    )
    config_loader = PDFConfigLoader(
        config_path=config.get("config_path")
    )

    totals_by_header: dict[str, Decimal] = defaultdict(
        lambda: Decimal("0")
    )
    processed_rows: list[dict[str, Any]] = []

    for source_file in source_files:
        print(f"Processing PDF: {source_file}")
        record = pdf_processor.extract_record(
            source_file,
            month=month,
        )

        print(
            f"Extracted: {record['location']} "
            f"{record['total_amount_tonnes']} tonnes"
        )

        location_config = (
            config_loader.get_location_config(
                record["location_key"]
            )
        )

        target_header = location_config.get(
            "target_header"
        )

        if not target_header:
            raise FoodWasteTransferError(
                "No target_header configured for "
                f"{record['location']}."
            )

        tonnes_text = record.get(
            "total_amount_tonnes"
        )

        if tonnes_text is None:
            raise FoodWasteTransferError(
                f"No total amount was found in {source_file}."
            )

        tonnes = Decimal(str(tonnes_text))
        kilograms = tonnes * Decimal("1000")

        totals_by_header[target_header] += kilograms

        processed_rows.append(
            {
                "source_file": str(source_file),
                "location_key": record["location_key"],
                "location": record["location"],
                "target_header": target_header,
                "month": month,
                "total_amount_tonnes": str(tonnes),
                "total_amount_kg": float(kilograms),
            }
        )

    return month, dict(totals_by_header), processed_rows


def _transfer_food_waste(
    config: dict[str, Any],
    *,
    preview: bool,
) -> dict[str, Any]:
    target_file = Path(config["target_file"])
    output_file = Path(config["target_output_file"])

    if not target_file.is_file():
        raise FileNotFoundError(
            f"Target workbook was not found: {target_file}"
        )

    output_path = _copy_workbook(
        target_file,
        output_file,
    )

    workbook = load_workbook(output_path)

    target_config = config.get("target", {})
    sheet_name = target_config.get("sheet_name")

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise FoodWasteTransferError(
                f"Worksheet '{sheet_name}' was not found."
            )

        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.active

    header_row = int(
        target_config.get("header_row", 1)
    )
    data_start_row = int(
        target_config.get(
            "data_start_row",
            header_row + 1,
        )
    )

    period_column_letter = target_config.get(
        "period_column",
        "A",
    )
    period_column = openpyxl.utils.column_index_from_string(
        period_column_letter
    )

    month, totals_by_header, extraction_rows = (
        _extract_pdf_values(config)
    )

    header_columns = _find_header_columns(
        worksheet,
        header_row,
    )

    target_row = _find_period_row(
        worksheet,
        period_column=period_column,
        data_start_row=data_start_row,
        month=month,
    )

    is_new_row = target_row is None

    if target_row is None:
        target_row = _next_period_row(
            worksheet,
            period_column=period_column,
            data_start_row=data_start_row,
        )

        previous_row = max(
            data_start_row,
            target_row - 1,
        )

        _copy_row_format(
            worksheet,
            previous_row,
            target_row,
        )

        worksheet.cell(
            row=target_row,
            column=period_column,
        ).value = month

    updated_values: dict[str, dict[str, Any]] = {}

    # Walk the target worksheet from left to right. This makes the
    # frontend preview follow the exact physical column sequence of
    # the uploaded target workbook, regardless of PDF upload order.
    totals_by_normalized_header = {
        _normalize_header(header): kilograms
        for header, kilograms in totals_by_header.items()
    }

    for column_number in range(1, worksheet.max_column + 1):
        target_header = worksheet.cell(
            row=header_row,
            column=column_number,
        ).value
        normalized_header = _normalize_header(target_header)

        if normalized_header not in totals_by_normalized_header:
            continue

        kilograms = totals_by_normalized_header[normalized_header]
        cell = worksheet.cell(
            row=target_row,
            column=column_number,
        )

        old_value = cell.value
        new_value = float(kilograms)
        cell.value = new_value

        if preview:
            cell.fill = PREVIEW_FILL

        updated_values[str(target_header)] = {
            "cell": cell.coordinate,
            "old_value": old_value,
            "new_value": new_value,
            "final_value": new_value,
            "status": (
                "new_value"
                if old_value in (None, "")
                else "updated"
            ),
        }

    missing_headers = [
        header
        for header in totals_by_header
        if _normalize_header(header) not in header_columns
    ]
    if missing_headers:
        raise FoodWasteTransferError(
            "Target column(s) were not found in the workbook: "
            + ", ".join(missing_headers)
        )

    # Calculated fields are intentionally excluded from updated_values,
    # so Total (kg) and Yearly Total (kg) do not appear in the Excel
    # Layout Preview. They are returned separately as processed_rows.
    calculation_config = dict(config)
    calculation_config["category"] = "food_waste"
    calculation_config["calculation_months"] = {
        "mode": "selected",
        "months": [month],
    }

    processed_rows = data_processing(
        workbook,
        calculation_config,
    )

    workbook.save(output_path)

    return {
        # Keep the normalized reporting month in the preview response.
        # ApplyPanel sends this value back to /apply; without it the apply
        # request receives an empty month and PDF extraction fails.
        "month": month,
        "updated_rows": [
            {
                "period": month,
                "target_row": target_row,
                "row_status": (
                    "added"
                    if is_new_row
                    else "existing"
                ),
                "is_new_row": is_new_row,
                "values": updated_values,
            }
        ],
        "processed_rows": processed_rows,
        "extraction_rows": extraction_rows,
        "inconsistencies": [],
        "new_columns": [],
        "skipped_columns": [],
        "rows_reviewed": 1,
        "output_file": str(output_path),
    }


def preview_food_waste_transfer(
    config: dict[str, Any],
) -> dict[str, Any]:
    return _transfer_food_waste(
        config,
        preview=True,
    )


def apply_food_waste_transfer(
    config: dict[str, Any],
) -> dict[str, Any]:
    return _transfer_food_waste(
        config,
        preview=False,
    )