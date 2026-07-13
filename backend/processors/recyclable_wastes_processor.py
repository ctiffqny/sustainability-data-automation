from pathlib import Path
import shutil

import openpyxl

from backend.core.period_utils import (
    normalize,
    normalize_period,
)
from backend.core.excel_utils import (
    get_headers,
    set_value,
    find_month_row,
    find_cell_by_text,
    find_first_header_row,
    keep_only_named_sheets,
)
from backend.processors.calculation_processor import (
    to_number,
    round_to_reporting,
    convert_to_kg,
    data_processing,
)


def extract_total_by_category(
    worksheet,
    category_name,
    total_header,
):
    """
    Find a recyclable-waste category in the source sheet and
    return its numeric value and source-cell coordinate.
    """
    category_cell = find_cell_by_text(
        worksheet,
        category_name,
    )
    total_cell = find_cell_by_text(
        worksheet,
        total_header,
    )

    if category_cell is None:
        raise ValueError(
            f"Could not find source category: {category_name}"
        )

    if total_cell is None:
        raise ValueError(
            f"Could not find total header: {total_header}"
        )

    value_cell = worksheet.cell(
        row=category_cell.row,
        column=total_cell.column,
    )

    return (
        to_number(value_cell.value),
        value_cell.coordinate,
    )


def transfer_raw_food_data(target_workbook, config):
    """
    Copy raw food-waste data between sheets in the target workbook.

    This data is written to the generated workbook, but it is not
    added as separate rows to the frontend preview table.
    """
    settings = config["raw_food_data"]

    source_worksheet = target_workbook[
        settings["source_sheet"]
    ]
    target_worksheet = target_workbook[
        settings["target_sheet"]
    ]

    source_header_row = find_first_header_row(
        source_worksheet,
        settings["source_month_column"],
    )

    target_header_row = 1

    source_headers = get_headers(
        source_worksheet,
        source_header_row,
    )
    target_headers = get_headers(
        target_worksheet,
        target_header_row,
    )

    source_month_column = source_headers.get(
        normalize(settings["source_month_column"])
    )
    source_value_column = source_headers.get(
        normalize(settings["source_value_column"])
    )
    target_value_column = target_headers.get(
        normalize(settings["target_value_column"])
    )

    if not source_month_column:
        raise ValueError(
            "Could not find raw-food source month column: "
            f"{settings['source_month_column']}"
        )

    if not source_value_column:
        raise ValueError(
            "Could not find raw-food source value column: "
            f"{settings['source_value_column']}"
        )

    if not target_value_column:
        raise ValueError(
            "Could not find raw-food target value column: "
            f"{settings['target_value_column']}"
        )

    for source_row in range(
        source_header_row + 1,
        source_worksheet.max_row + 1,
    ):
        month = source_worksheet.cell(
            row=source_row,
            column=source_month_column,
        ).value

        if month in ("", None):
            continue

        value = source_worksheet.cell(
            row=source_row,
            column=source_value_column,
        ).value

        if value in ("", None):
            continue

        target_row = find_month_row(
            target_worksheet,
            month,
        )

        target_worksheet.cell(
            row=target_row,
            column=target_value_column,
        ).value = value


def generate_recyclable_waste_output(
    config,
    output_path,
):
    """
    Generate and save the completed recyclable-waste target copy.

    Returns metadata needed to build the preview from the saved copy.
    """
    skipped_columns = set()
    inconsistencies = []

    source_workbook = openpyxl.load_workbook(
        config["source_file"],
        data_only=True,
    )

    # Do not use data_only=True here because target formulas need
    # to remain formulas in the generated workbook.
    target_workbook = openpyxl.load_workbook(
        config["target_file"],
    )

    source_worksheet = source_workbook[
        config["source_sheet"]
    ]
    target_worksheet = target_workbook[
        config["target_sheet"]
    ]

    target_header_row = 1
    target_headers = get_headers(
        target_worksheet,
        target_header_row,
    )

    # MAY-26 in the current YAML configuration.
    period_value = config["source_sheet"]

    target_row = find_month_row(
        target_worksheet,
        period_value,
    )

    period_column_name = config.get(
        "target_period_column",
        "Period",
    )

    set_value(
        target_worksheet,
        target_row,
        target_headers,
        period_column_name,
        period_value,
        skipped_columns,
    )

    source_cells_by_column = {}

    for target_column, source_categories in (
        config["column_map"].items()
    ):
        categories = (
            source_categories
            if isinstance(source_categories, list)
            else [source_categories]
        )

        source_value = 0
        source_cells = []

        for source_category in categories:
            category_value, source_cell = (
                extract_total_by_category(
                    source_worksheet,
                    source_category,
                    config["source_total_header"],
                )
            )

            converted_value = convert_to_kg(
                source_category,
                category_value,
            )

            source_value += converted_value
            source_cells.append(source_cell)

        source_cells_by_column[target_column] = (
            source_cells
        )

        target_column_number = target_headers.get(
            normalize(target_column)
        )

        if not target_column_number:
            skipped_columns.add(target_column)
            continue

        target_cell = target_worksheet.cell(
            row=target_row,
            column=target_column_number,
        )

        existing_target_value = target_cell.value

        # Preserve existing formulas.
        if (
            isinstance(existing_target_value, str)
            and existing_target_value.startswith("=")
        ):
            continue

        if source_value in ("", None):
            continue

        target_reported = round_to_reporting(
            existing_target_value
        )
        source_reported = round_to_reporting(
            source_value
        )

        if (
            existing_target_value not in ("", None)
            and target_reported != source_reported
        ):
            inconsistencies.append(
                {
                    "period": normalize_period(
                        period_value
                    ),
                    "column": target_column,
                    "existing_target_value": (
                        target_reported
                    ),
                    "new_source_value": (
                        source_reported
                    ),
                    "target_cell": (
                        target_cell.coordinate
                    ),
                    "source_cell": ", ".join(
                        source_cells
                    ),
                }
            )

            # Preserve the existing target value.
            continue

        target_cell.value = source_value

    transfer_raw_food_data(
        target_workbook,
        config,
    )

    data_processing(
        target_workbook,
        config,
    )

    keep_only_named_sheets(
        target_workbook,
        [
            config["target_sheet"],
            config["calculation_target_sheet"],
        ],
    )

    output_path = Path(output_path)
    output_path.parent.mkdir(
        parents=True,
        exist_ok=True,
    )

    target_workbook.save(output_path)

    return {
        "output_path": str(output_path),
        "target_row": target_row,
        "period_value": period_value,
        "skipped_columns": sorted(
            skipped_columns
        ),
        "inconsistencies": inconsistencies,
        "source_cells_by_column": (
            source_cells_by_column
        ),
    }


def build_recyclable_waste_preview(
    generated_file,
    config,
    generation_result,
):
    """
    Reopen the generated target copy and build the React preview
    from the values actually saved in that workbook.
    """
    workbook = openpyxl.load_workbook(
        generated_file,
        data_only=False,
    )

    worksheet = workbook[
        config["target_sheet"]
    ]

    header_row = 1
    headers = get_headers(
        worksheet,
        header_row,
    )

    target_row = generation_result["target_row"]

    period_column_name = config.get(
        "target_period_column",
        "Period",
    )

    period_column_number = headers.get(
        normalize(period_column_name)
    )

    if not period_column_number:
        raise ValueError(
            "Could not find target period column in "
            f"generated workbook: {period_column_name}"
        )

    period_value = worksheet.cell(
        row=target_row,
        column=period_column_number,
    ).value

    inconsistency_cells = {
        item["target_cell"]
        for item in generation_result[
            "inconsistencies"
        ]
    }

    values = {}

    for column_name in config["column_map"]:
        column_number = headers.get(
            normalize(column_name)
        )

        if not column_number:
            continue

        cell = worksheet.cell(
            row=target_row,
            column=column_number,
        )

        if cell.coordinate in inconsistency_cells:
            status = (
                "inconsistency_not_overwritten"
            )
        elif (
            isinstance(cell.value, str)
            and cell.value.startswith("=")
        ):
            status = "formula_preserved"
        else:
            status = "generated_output"

        source_cells = generation_result[
            "source_cells_by_column"
        ].get(column_name, [])

        values[column_name] = {
            "cell": cell.coordinate,
            "source_cell": ", ".join(
                source_cells
            ),
            "status": status,
            "old_value": None,
            "new_value": cell.value,
            "final_value": cell.value,
        }

    return [
        {
            "period": period_value,
            "target_row": target_row,
            "is_new_row": False,
            "row_status": "generated",
            "values": values,
        }
    ]


def run_recyclable_wastes_transfer(
    config,
    save_outputs,
):
    """
    Generate the target copy first, then read that copy to produce
    the frontend preview response.
    """
    generation_result = (
        generate_recyclable_waste_output(
            config=config,
            output_path=config[
                "target_output_file"
            ],
        )
    )

    updated_rows = (
        build_recyclable_waste_preview(
            generated_file=generation_result[
                "output_path"
            ],
            config=config,
            generation_result=generation_result,
        )
    )

    source_output_file = config.get(
        "source_output_file"
    )

    if save_outputs and source_output_file:
        source_output_path = Path(
            source_output_file
        )
        source_output_path.parent.mkdir(
            parents=True,
            exist_ok=True,
        )

        # # There is currently no waste-specific source
        # # highlighting, so save an unchanged source copy.
        # shutil.copy2(
        #     config["source_file"],
        #     source_output_path,
        # )

    messages = [
        (
            "Generated the recyclable-waste target "
            "workbook and built the preview from "
            "the saved copy."
        )
    ]

    if generation_result["skipped_columns"]:
        messages.append(
            "Some target columns were skipped."
        )
    else:
        messages.append(
            "No columns were skipped."
        )

    if generation_result["inconsistencies"]:
        messages.append(
            "Inconsistent target values were preserved "
            "and were not overwritten."
        )
    else:
        messages.append(
            "No inconsistencies were found."
        )

    return {
        "updated_rows": updated_rows,
        "new_columns": [],
        "skipped_columns": generation_result[
            "skipped_columns"
        ],
        "inconsistencies": generation_result[
            "inconsistencies"
        ],
        "messages": messages,
        "saved": save_outputs,
        "preview_file": generation_result[
            "output_path"
        ],
    }


def preview_recyclable_wastes_transfer(config):
    # Preview intentionally generates and saves a temporary
    # target copy, because the preview is read from that copy.
    return run_recyclable_wastes_transfer(
        config,
        save_outputs=False,
    )


def apply_recyclable_wastes_transfer(
    config,
    output_mode="duplicate",
):
    if output_mode not in {
        "duplicate",
        "amend",
    }:
        raise ValueError(
            "output_mode must be "
            "'duplicate' or 'amend'"
        )

    return run_recyclable_wastes_transfer(
        config,
        save_outputs=True,
    )


def process_recyclable_wastes(
    config,
    preview_mode=False,
):
    result = run_recyclable_wastes_transfer(
        config,
        save_outputs=not preview_mode,
    )

    for message in result["messages"]:
        print(message)

    return result