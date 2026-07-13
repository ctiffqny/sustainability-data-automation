from copy import copy
from openpyxl.utils import get_column_letter


def copy_cell_format(source_cell, target_cell):
    if source_cell.has_style:
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)


def copy_column_format_from_source(
    source_ws,
    target_ws,
    source_col,
    target_col,
    source_header_row,
    target_header_row,
    source_last_row=None,
):
    target_ws.column_dimensions[get_column_letter(target_col)].width = (
        source_ws.column_dimensions[get_column_letter(source_col)].width
    )

    row_offset = target_header_row - source_header_row

    if source_last_row is None:
        source_last_row = source_ws.max_row

    for source_row in range(1, source_last_row + 1):
        target_row = source_row + row_offset

        if target_row < 1 or target_row > target_ws.max_row:
            continue

        copy_cell_format(
            source_ws.cell(row=source_row, column=source_col),
            target_ws.cell(row=target_row, column=target_col),
        )

def copy_placeholder_formats_from_target(
    target_ws,
    template_col,
    target_col,
    period_col,
    first_placeholder_row,
):
    from backend.core.period_utils import parse_period

    for target_row in range(first_placeholder_row, target_ws.max_row + 1):
        period_value = target_ws.cell(
            row=target_row,
            column=period_col,
        ).value

        if parse_period(period_value) is None:
            continue

        copy_cell_format(
            target_ws.cell(row=target_row, column=template_col),
            target_ws.cell(row=target_row, column=target_col),
        )


def copy_row_format(source_ws, target_ws, source_row, target_row):
    for col in range(1, target_ws.max_column + 1):
        copy_cell_format(
            source_ws.cell(row=source_row, column=col),
            target_ws.cell(row=target_row, column=col),
        )