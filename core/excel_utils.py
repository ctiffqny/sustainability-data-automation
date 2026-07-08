from openpyxl.styles import PatternFill

from core.period_utils import (
    normalize,
    parse_period,
    normalize_period,
    periods_match,
    first_day_of_month,
)


DISCREPANCY_FILL = PatternFill(
    fill_type = "solid",
    start_color="FFFF00",
    end_color="FFFF00"
)


def find_first_header_row(ws, required_header="Period"):
    required = normalize(required_header)

    for row in ws.iter_rows():
        values = [normalize(cell.value) for cell in row]
        if required in values:
            return row[0].row

    raise ValueError(f"Could not find first header row containing {required_header}")


def get_headers(ws, header_row):
    headers = {}

    for cell in ws[header_row]:
        if cell.value:
            headers[normalize(cell.value)] = cell.column

    return headers


def get_value(ws, row_num, headers, header_name, skipped_columns):
    col = headers.get(normalize(header_name))

    if not col:
        skipped_columns.add(header_name)
        return ""

    value = ws.cell(row=row_num, column=col).value
    return "" if value is None else value


def set_value(ws, row_num, headers, header_name, value, skipped_columns):
    col = headers.get(normalize(header_name))

    if not col:
        skipped_columns.add(header_name)
        return

    ws.cell(row=row_num, column=col).value = value


def find_period_row(ws, headers, period_value):
    period_col = headers.get(normalize("Period"))

    if not period_col:
        raise ValueError("Could not find Period column in target file")

    for row in range(1, ws.max_row + 1):
        if periods_match(ws.cell(row=row, column=period_col).value, period_value):
            return row

    return ws.max_row + 1   # add new row if period not found


def find_month_row(target_ws, month):
    wanted = normalize_period(month)

    for row in range(2, target_ws.max_row + 1):
        value = target_ws.cell(row=row, column=1).value

        if normalize_period(value) == wanted:
            return row

    raise ValueError(f"Could not find month {month} in column A")


# highlight inconsistency in target

def flag_inconsistency(
    target_ws,
    source_ws_highlight,
    target_row,
    target_col,
    source_row,
    source_col,
    inconsistencies,
    period_value,
    column_name,
    existing_value,
    new_value,
    second_source_col=None
):
    target_cell = target_ws.cell(row=target_row, column=target_col)
    source_cell = source_ws_highlight.cell(row=source_row, column=source_col)

    target_cell.fill = DISCREPANCY_FILL
    source_cell.fill = DISCREPANCY_FILL

    source_cells = [source_cell.coordinate]

    if second_source_col is not None:
        second_source_cell = source_ws_highlight.cell(row=source_row, column=second_source_col)
        second_source_cell.fill = DISCREPANCY_FILL
        source_cells.append(second_source_cell.coordinate)

    inconsistencies.append({
        "period": normalize_period(period_value),
        "column": column_name,
        "existing_target_value": existing_value,
        "new_source_value": new_value,
        "target_cell": target_cell.coordinate,
        "source_cell": source_cell.coordinate
    })


def find_cell_by_text(ws, text):
    wanted = normalize(text)

    for row in ws.iter_rows():
        for cell in row:
            if normalize(cell.value) == wanted:
                return cell

    return None


def keep_only_relevant_sheets(wb, target_sheet):
    ws = wb[target_sheet]
    sheets_to_keep = {target_sheet}

    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.startswith("="):
                for sheet_name in wb.sheetnames:
                    if (
                        f"'{sheet_name}'!" in cell.value
                        or f"{sheet_name}!" in cell.value
                    ):
                        sheets_to_keep.add(sheet_name)

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]


def keep_only_named_sheets(wb, sheet_names_to_keep):
    sheets_to_keep = set(sheet_names_to_keep)

    for sheet_name in list(wb.sheetnames):
        if sheet_name not in sheets_to_keep:
            del wb[sheet_name]