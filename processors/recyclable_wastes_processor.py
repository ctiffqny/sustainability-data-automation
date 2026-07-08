import openpyxl

from core.period_utils import normalize, normalize_period
from core.excel_utils import (
    get_headers,
    set_value,
    find_month_row,
    find_cell_by_text,
    keep_only_relevant_sheets,
)

from processors.calculation_processor import (
    to_number,
    round_to_reporting,
    convert_to_kg,
)


def extract_total_by_category(ws, category_name, total_header):
    category_cell = find_cell_by_text(ws, category_name)
    total_cell = find_cell_by_text(ws, total_header)

    if category_cell is None:
        raise ValueError(f"Could not find source category: {category_name}")

    if total_cell is None:
        raise ValueError(f"Could not find total header: {total_header}")

    value = ws.cell(
        row=category_cell.row,
        column=total_cell.column
    ).value

    return to_number(value)


# INCONSISTENCY TESTING FOR WASTE

def get_recent_source_sheets(source_wb, current_sheet_name, lookback_count=12):
    sheet_names = source_wb.sheetnames

    if current_sheet_name not in sheet_names:
        raise ValueError(f"Could not find source sheet: {current_sheet_name}")

    current_index = sheet_names.index(current_sheet_name)

    start_index = max(0, current_index - lookback_count + 1)

    return sheet_names[start_index: current_index + 1]


def transfer_recyclable_wastes(config):
    skipped_columns = []
    inconsistencies = []

    source_wb = openpyxl.load_workbook(config["source_file"], data_only=True)
    target_wb = openpyxl.load_workbook(config["target_file"])

    source_ws = source_wb[config["source_sheet"]]
    target_ws = target_wb[config["target_sheet"]]

    target_header_row = 1
    target_headers = get_headers(target_ws, target_header_row)

    period_value = config["source_sheet"]
    target_row = find_month_row(target_ws, period_value)

    set_value(
        target_ws,
        target_row,
        target_headers,
        "Period",
        period_value,
        set()
    )

    for target_col, source_categories in config["column_map"].items():

        if isinstance(source_categories, list):
            source_value = sum(
                convert_to_kg(
                    source_category,
                    extract_total_by_category(
                        source_ws,
                        source_category,
                        config["source_total_header"]
                    )
                )
                for source_category in source_categories
            )

        else:
            source_value = extract_total_by_category(
                source_ws,
                source_categories,
                config["source_total_header"]
            )
        
        if not isinstance(source_categories, list):
            source_value = convert_to_kg(source_categories, source_value)

        target_col_num = target_headers.get(normalize(target_col))

        if not target_col_num:
            skipped_columns.append(target_col)
            continue

        target_cell = target_ws.cell(row=target_row, column=target_col_num)

        # preserve any existing formula in the target cell, for any category
        if isinstance(target_cell.value, str) and target_cell.value.startswith("="):
            continue

        # do not overwrite with blank source values
        if source_value in ("", None):
            continue

        existing_target_value = target_cell.value

        target_reported = round_to_reporting(existing_target_value)
        source_reported = round_to_reporting(source_value)

        if (
            existing_target_value not in ("", None)
            and target_reported != source_reported
        ):
            inconsistencies.append({
                "period": normalize_period(period_value),
                "column": target_col,
                "existing_target_value": target_reported,
                "new_source_value": source_reported,
                "target_cell": target_cell.coordinate,
            })

            continue

        target_cell.value = source_value

    keep_only_relevant_sheets(target_wb, config["target_sheet"])
    target_wb.save(config["target_output_file"])

    print(f"\nSuccess. Saved as {config['target_output_file']}")
    print("Inconsistencies were reported only. Cell highlighting is currently disabled.")

    if skipped_columns:
        print("\nMissing target columns:")
        for col in skipped_columns:
            print(f"- {col}")
    else:
        print("\nNo columns were skipped.")

    if inconsistencies:
        print("\n=============")
        print("\nNOTE:")
        print("Cells with inconsistencies were NOT overwritten.")
        print("The existing values in the target workbook were preserved.")

        print("\nInconsistencies found:")

        print("\nSUST office reference - output workbook cell numbers:")
        for item in inconsistencies:
            print(
                f"- {item['period']} | {item['column']} ({item['target_cell']}): "
                f"SUST had {item['existing_target_value']}, "
                f"source has {item['new_source_value']}"
            )

    else:
        print("\nNo inconsistencies found.")


def process_recyclable_wastes(config):
    transfer_recyclable_wastes(config)