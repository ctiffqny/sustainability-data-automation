import openpyxl

from core.period_utils import (
    normalize,
    parse_period,
    first_day_of_month,
    subtract_months,
)

from core.excel_utils import (
    find_first_header_row,
    get_headers,
    get_value,
    find_period_row,
    flag_inconsistency,
    keep_only_relevant_sheets,
)

from processors.calculation_processor import (
    resolve_cell_value,
    values_different,
    numeric_difference,
)

from core.excel_format_utils import (
    copy_cell_format,
    copy_column_format_from_source,
)

from core.excel_layout_utils import (
    find_previous_source_header_in_target,
    ensure_source_separators_exist,
    copy_source_separator_after_if_exists,
)


def format_new_period_row(target_ws, target_row, period_col, period_date):
    """
    Copies formatting from the previous row and forces Period format to May-26 style.
    """
    template_row = target_row - 1

    if template_row >= 1:
        for col in range(1, target_ws.max_column + 1):
            copy_cell_format(
                target_ws.cell(row=template_row, column=col),
                target_ws.cell(row=target_row, column=col),
            )

    period_cell = target_ws.cell(row=target_row, column=period_col)
    period_cell.value = first_day_of_month(period_date)
    period_cell.number_format = "mmm-yy"


def add_new_source_columns_to_output(
    source_ws,
    target_ws,
    source_header_row,
    target_header_row,
    source_headers,
    target_headers,
    start_date,
    end_date,
    config,
):
    new_columns = []

    mapped_source_columns = {
        normalize(source_col)
        for source_col in config["column_map"].values()
    }
    mapped_source_columns.add(normalize("Period"))

    for cell in source_ws[source_header_row]:
        if not cell.value:
            continue

        source_header_name = str(cell.value).strip()
        normalized_source_name = normalize(source_header_name)

        if normalized_source_name in mapped_source_columns:
            continue

        source_col_num = cell.column

        if normalized_source_name not in target_headers:
            previous_source_col, previous_target_col = find_previous_source_header_in_target(
                source_ws,
                source_header_row,
                source_col_num,
                target_headers,
            )

            if previous_target_col is None:
                raise ValueError(
                    f"Could not find where to insert new column: {source_header_name}"
                )

            new_col_num = previous_target_col + 1

            # Copy any blank separator columns that exist in the source
            for source_gap_col in range(previous_source_col + 1, source_col_num):
                source_gap_value = source_ws.cell(
                    row=source_header_row,
                    column=source_gap_col,
                ).value

                if source_gap_value not in ("", None):
                    continue

                if target_ws.cell(
                    row=target_header_row,
                    column=new_col_num,
                ).value not in ("", None):
                    target_ws.insert_cols(new_col_num)

                copy_column_format_from_source(
                    source_ws,
                    target_ws,
                    source_gap_col,
                    new_col_num,
                    source_header_row,
                    target_header_row,
                )

                new_col_num += 1

            # Insert new building column
            target_ws.insert_cols(new_col_num)

            copy_column_format_from_source(
                source_ws,
                target_ws,
                source_col_num,
                new_col_num,
                source_header_row,
                target_header_row,
            )

            target_ws.cell(
                row=target_header_row,
                column=new_col_num,
            ).value = source_header_name

            # Copy blank separator immediately after this source building, if source has one
            source_separator_col = source_col_num + 1

            if (
                source_separator_col <= source_ws.max_column
                and source_ws.cell(
                    row=source_header_row,
                    column=source_separator_col,
                ).value in ("", None)
            ):
                target_separator_col = new_col_num + 1

                if target_ws.cell(
                    row=target_header_row,
                    column=target_separator_col,
                ).value not in ("", None):
                    target_ws.insert_cols(target_separator_col)

                copy_column_format_from_source(
                    source_ws,
                    target_ws,
                    source_separator_col,
                    target_separator_col,
                    source_header_row,
                    target_header_row,
                )

            target_headers.clear()
            target_headers.update(get_headers(target_ws, target_header_row))

            config["column_map"][source_header_name] = source_header_name
            new_columns.append(source_header_name)

        else:
            new_col_num = target_headers[normalized_source_name]
            config["column_map"][source_header_name] = source_header_name

        period_col = target_headers.get(normalize("Period"))

        if not period_col:
            raise ValueError("Could not find Period column after inserting new column")

        for row in range(target_header_row + 1, target_ws.max_row + 1):
            period_value = target_ws.cell(row=row, column=period_col).value
            period_date = parse_period(period_value)

            if period_date is None:
                continue

            period_month = first_day_of_month(period_date)

            if period_month <= end_date:
                if target_ws.cell(row=row, column=new_col_num).value in ("", None):
                    target_ws.cell(row=row, column=new_col_num).value = 0

    return new_columns


def process_electricity(config):
    skipped_columns = set()
    inconsistencies = []

    source_wb = openpyxl.load_workbook(config["source_file"], data_only=True)
    source_wb_highlight = openpyxl.load_workbook(config["source_file"], data_only=True)
    target_wb = openpyxl.load_workbook(config["target_file"])

    source_ws = source_wb[config["source_sheet"]]
    source_ws_highlight = source_wb_highlight[config["source_sheet"]]
    target_ws = target_wb[config["target_sheet"]]

    source_header_row = find_first_header_row(source_ws, "Period")
    target_header_row = find_first_header_row(target_ws, "Period")

    source_headers = get_headers(source_ws, source_header_row)
    target_headers = get_headers(target_ws, target_header_row)

    source_period_col = source_headers.get(normalize("Period"))

    if not source_period_col:
        raise ValueError("Could not find Period column in source file")

    source_months = []

    for source_row in range(source_header_row + 1, source_ws.max_row + 1):
        period_value = source_ws.cell(row=source_row, column=source_period_col).value
        period_date = parse_period(period_value)

        if period_date is not None:
            source_months.append(first_day_of_month(period_date))

    if not source_months:
        raise ValueError("No valid periods found in source file")

    end_date = max(source_months)
    start_date = subtract_months(end_date, config["lookback_months"])

    new_columns = add_new_source_columns_to_output(
        source_ws,
        target_ws,
        source_header_row,
        target_header_row,
        source_headers,
        target_headers,
        start_date,
        end_date,
        config,
    )

    for source_row in range(source_header_row + 1, source_ws.max_row + 1):
        period_value = source_ws.cell(row=source_row, column=source_period_col).value
        period_date = parse_period(period_value)

        if period_date is None:
            continue

        period_month = first_day_of_month(period_date)

        if not (start_date <= period_month <= end_date):
            continue

        target_row = find_period_row(target_ws, target_headers, period_value)

        period_col = target_headers.get(normalize("Period"))

        if not period_col:
            raise ValueError("Could not find Period column in target file")

        if target_row == target_ws.max_row + 1:
            format_new_period_row(
                target_ws,
                target_row,
                period_col,
                period_date,
            )
        else:
            period_cell = target_ws.cell(row=target_row, column=period_col)
            period_cell.value = first_day_of_month(period_date)
            period_cell.number_format = "mmm-yy"

        for target_col, source_col in config["column_map"].items():
            source_value = get_value(
                source_ws,
                source_row,
                source_headers,
                source_col,
                skipped_columns,
            )

            target_col_num = target_headers.get(normalize(target_col))

            if not target_col_num:
                skipped_columns.add(target_col)
                continue

            existing_target_value = resolve_cell_value(
                target_ws,
                target_ws.cell(row=target_row, column=target_col_num).value,
            )

            if (
                target_col not in new_columns
                and existing_target_value not in ("", None)
                and source_value not in ("", None)
                and values_different(existing_target_value, source_value)
            ):
                diff = numeric_difference(existing_target_value, source_value)

                if diff is None or diff >= 1:
                    flag_inconsistency(
                        target_ws,
                        source_ws_highlight,
                        target_row,
                        target_col_num,
                        source_row,
                        source_headers.get(normalize(source_col)),
                        inconsistencies,
                        period_value,
                        target_col,
                        existing_target_value,
                        source_value,
                    )
                    continue

            if source_value in ("", None):
                pass
            else:
                target_ws.cell(row=target_row, column=target_col_num).value = source_value

    keep_only_relevant_sheets(target_wb, config["target_sheet"])
    target_wb.save(config["target_output_file"])

    source_wb_highlight.save(config["source_output_file"])

    print(f"\nSuccess. Saved as {config['target_output_file']}")
    print(f"Highlighted source copy saved as {config['source_output_file']}")

    print(
        f"This application compared and transferred data only within the "
        f"last {config['lookback_months']} months "
        f"({start_date.strftime('%b %y')} to {end_date.strftime('%b %y')})."
    )

    if new_columns:
        print("\nNew source columns detected and added to output:")
        for col in new_columns:
            print(f"- {col}")
            print(
                "Previous months were filled with 0, and available source data "
                "has already been updated in the output Excel file."
            )

    if skipped_columns:
        print("\nMissing columns (possibly due to naming error):")
        for col in sorted(skipped_columns):
            print(f"- {col}")
    else:
        print("\nNo columns were skipped.")

    if inconsistencies:
        print("\n=============")
        print("\nNOTE:")
        print("Cells with inconsistencies were NOT overwritten.")
        print("The existing values in the target workbook were preserved.")

        print("\nInconsistencies found:")

        print("\nSUST office reference - output workbook cell numbers:")
        for item in inconsistencies:
            print(
                f"- {item['period']} | {item['column']} ({item['target_cell']}): "
                f"SUST had {item['existing_target_value']}, "
                f"source has {item['new_source_value']}"
            )

        print("\nOther department reference - source workbook cell numbers:")
        for item in inconsistencies:
            print(
                f"- {item['period']} | {item['column']} ({item['source_cell']}): "
                f"SUST had {item['existing_target_value']}, "
                f"source has {item['new_source_value']}"
            )

    else:
        print("\nNo inconsistencies found.")