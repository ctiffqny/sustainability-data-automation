import openpyxl
import re
from datetime import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import operator
from openpyxl.styles import PatternFill
from copy import copy

def run_excel_transfer(config):

    DISCREPANCY_FILL = PatternFill(
        fill_type = "solid",
        start_color="FFFF00",
        end_color="FFFF00"
    )

    def save_changed_sheet_only(target_wb, sheet_name, output_file):
        source_ws = target_wb[sheet_name]

        output_wb = openpyxl.Workbook()
        output_ws = output_wb.active
        output_ws.title = sheet_name

        for row in source_ws.iter_rows():
            for cell in row:
                new_cell = output_ws[cell.coordinate]
                new_cell.value = cell.value

                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.fill = copy(cell.fill)
                    new_cell.border = copy(cell.border)
                    new_cell.alignment = copy(cell.alignment)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)

        for col_letter, dimension in source_ws.column_dimensions.items():
            output_ws.column_dimensions[col_letter].width = dimension.width

        for row_num, dimension in source_ws.row_dimensions.items():
            output_ws.row_dimensions[row_num].height = dimension.height

        for merged_range in source_ws.merged_cells.ranges:
            output_ws.merge_cells(str(merged_range))

        output_wb.save(output_file)

    # system calculates excel formulas before comparing
    OPERATORS = {
        "+": operator.add,
        "-": operator.sub,
        "*": operator.mul,
        "/": operator.truediv,
    }


    def resolve_cell_value(ws, value):
        if isinstance(value, str) and value.startswith("="):
            return calculate_simple_formula(ws, value)

        return value


    def calculate_simple_formula(ws, formula):
        original_formula = formula

        formula = formula.strip().lstrip("=")

        # handles number related formulas in excel
        if re.fullmatch(r"\s*-?\d+(?:\.\d+)?(?:\s*[+\-*/]\s*-?\d+(?:\.\d+)?)+\s*", formula):
            try:
                return eval(formula, {"__builtins__": {}})
            except Exception:
                return original_formula

        # handles position related formulas in excel e.g. =Y146-Z146
        match = re.fullmatch(r"\s*([A-Z]+\d+)\s*([+\-*/])\s*([A-Z]+\d+)\s*", formula)
        if match:
            left_cell, op, right_cell = match.groups()

            left_value = resolve_cell_value(ws, ws[left_cell].value)
            right_value = resolve_cell_value(ws, ws[right_cell].value)

            try:
                left_value = float(left_value)
                right_value = float(right_value)

                if op == "+":
                    return left_value + right_value
                if op == "-":
                    return left_value - right_value
                if op == "*":
                    return left_value * right_value
                if op == "/":
                    return left_value / right_value
            except Exception:
                return original_formula

        return original_formula

    # strip data of everything so formatting doesn't affect

    def normalize(text):
        if text is None:
            return ""
        text = str(text).lower().strip()
        text = re.sub(r"[^a-z0-9]+", "", text)
        return text

    # parse and match rows by period (just a safety measure in case the formatting is ever changed)

    def parse_period(value):
        if value is None:
            return None

        if isinstance(value, datetime):
            return value

        text = str(value).strip().replace("Sept", "Sep")

        for fmt in ("%b-%y", "%b %y", "%Y-%m-%d", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt)
            except ValueError:
                pass

        return None
    
    def to_number(value):
        if value in ("", None):
            return 0

        try:
            return float(value)
        except (ValueError, TypeError):
            return 0
        
    def convert_to_kg(category_name, value):

        if value in ("", None):
            return 0

        value = to_number(value)

        tonne_categories = {
            normalize("General Waste"),
            normalize("General Wastes"),
            normalize("Trash"),
            normalize("Yellow Skip"),
            normalize("Yellow skip"),
        }

        if normalize(category_name) in tonne_categories:
            return value * 1000

        return value

    def normalize_period(value):
        parsed = parse_period(value)
        if parsed:
            return parsed.strftime("%b-%y").lower()
        return str(value).strip().lower()


    def periods_match(a, b):
        return normalize_period(a) == normalize_period(b)


    def first_day_of_month(date_value):
        return datetime(date_value.year, date_value.month, 1)


    def subtract_months(date_value, months):
        year = date_value.year
        month = date_value.month

        for _ in range(months):
            if month == 1:
                year -= 1
                month = 12
            else:
                month -= 1

        return datetime(year, month, 1)


    def find_first_header_row(ws, required_header="Period"):
        required = normalize(required_header)

        for row in ws.iter_rows():
            values = [normalize(cell.value) for cell in row]
            if required in values:
                return row[0].row

        raise ValueError(f"Could not find first header row containing {required_header}")


    def get_headers(ws, header_row):
        headers = {}

        for cell in ws[header_row]:
            if cell.value:
                headers[normalize(cell.value)] = cell.column

        return headers


    def get_value(ws, row_num, headers, header_name, skipped_columns):
        col = headers.get(normalize(header_name))

        if not col:
            skipped_columns.add(header_name)
            return ""

        value = ws.cell(row=row_num, column=col).value
        return "" if value is None else value


    def set_value(ws, row_num, headers, header_name, value, skipped_columns):
        col = headers.get(normalize(header_name))

        if not col:
            skipped_columns.add(header_name)
            return

        ws.cell(row=row_num, column=col).value = value


    def find_period_row(ws, headers, period_value):
        period_col = headers.get(normalize("Period"))

        if not period_col:
            raise ValueError("Could not find Period column in target file")

        for row in range(1, ws.max_row + 1):
            if periods_match(ws.cell(row=row, column=period_col).value, period_value):
                return row

        return ws.max_row + 1   # add new row if period not found
    

    def find_month_row(target_ws, month):
        wanted = normalize_period(month)

        for row in range(2, target_ws.max_row + 1):
            value = target_ws.cell(row=row, column=1).value

            if normalize_period(value) == wanted:
                return row

        raise ValueError(f"Could not find month {month} in column A")

    # compare differences in value

    def values_different(a, b):
        if a in ("", None) and b in ("", None):
            return False

        try:
            return abs(float(a) - float(b)) > 0.01
        except (ValueError, TypeError):
            return str(a).strip() != str(b).strip()
        
    def numeric_difference(a, b):
        try:
            return abs(float(a) - float(b))
        except (ValueError, TypeError):
            return None

    # highlight inconsistency in target

    def flag_inconsistency(
        target_ws,
        source_ws_highlight,
        target_row,
        target_col,
        source_row,
        source_col,
        inconsistencies,
        period_value,
        column_name,
        existing_value,
        new_value,
        second_source_col=None
    ):
        target_cell = target_ws.cell(row=target_row, column=target_col)
        source_cell = source_ws_highlight.cell(row=source_row, column=source_col)

        target_cell.fill = DISCREPANCY_FILL
        source_cell.fill = DISCREPANCY_FILL

        source_cells = [source_cell.coordinate]

        if second_source_col is not None:
            second_source_cell = source_ws_highlight.cell(row=source_row, column=second_source_col)
            second_source_cell.fill = DISCREPANCY_FILL
            source_cells.append(second_source_cell.coordinate)

        inconsistencies.append({
            "period": normalize_period(period_value),
            "column": column_name,
            "existing_target_value": existing_value,
            "new_source_value": new_value,
            "target_cell": target_cell.coordinate,
            "source_cell": source_cell.coordinate
        })

    # make sure new building is added after all previous buildings instead of end of sheet
    def find_last_building_column(headers):
        last_col = 0

        for target_col in config["column_map"].keys():
            col = headers.get(normalize(target_col))
            if col and col > last_col:
                last_col = col

        return last_col

    # new buildings may be added to Excel. automatically update new column to target
    def add_new_source_columns_to_output(
        source_ws,
        target_ws,
        source_header_row,
        target_header_row,
        source_headers,
        target_headers,
        start_date,
        end_date
    ):
        new_columns = []

        mapped_source_columns = {normalize(source_col) for source_col in config["column_map"].values()}
        mapped_source_columns.add(normalize("Period"))

        for cell in source_ws[source_header_row]:
            if not cell.value:
                continue

            source_header_name = str(cell.value).strip()
            normalized_source_name = normalize(source_header_name)

            if normalized_source_name in mapped_source_columns:
                continue

            if normalized_source_name not in target_headers:
                new_col_num = find_last_building_column(target_headers) + 1

                target_ws.insert_cols(new_col_num)
                target_ws.cell(row=target_header_row, column=new_col_num).value = source_header_name

                target_headers.clear()
                target_headers.update(get_headers(target_ws, target_header_row))

                config["column_map"][source_header_name] = source_header_name
                new_columns.append(source_header_name)
            else:
                new_col_num = target_headers[normalized_source_name]
                config["column_map"][source_header_name] = source_header_name

            period_col = target_headers.get(normalize("Period"))

            if not period_col:
                raise ValueError("Could not find Period column after inserting new column")

            for row in range(target_header_row + 1, target_ws.max_row + 1):
                period_value = target_ws.cell(row=row, column=period_col).value
                period_date = parse_period(period_value)

                if period_date is None:
                    continue

                period_month = first_day_of_month(period_date)

                if period_month <= end_date:
                    if target_ws.cell(row=row, column=new_col_num).value in ("", None):
                        target_ws.cell(row=row, column=new_col_num).value = 0

        return new_columns
        
    def find_cell_by_text(ws, text):
        wanted = normalize(text)

        for row in ws.iter_rows():
            for cell in row:
                if normalize(cell.value) == wanted:
                    return cell

        return None


    def extract_total_by_category(ws, category_name, total_header):
        category_cell = find_cell_by_text(ws, category_name)
        total_cell = find_cell_by_text(ws, total_header)

        if category_cell is None:
            raise ValueError(f"Could not find source category: {category_name}")

        if total_cell is None:
            raise ValueError(f"Could not find total header: {total_header}")

        value = ws.cell(
            row=category_cell.row,
            column=total_cell.column
        ).value

        return to_number(value)


    def transfer_recent_months(config):
        skipped_columns = set()
        inconsistencies = []

        source_wb = openpyxl.load_workbook(config["source_file"], data_only=True)
        source_wb_highlight = openpyxl.load_workbook(config["source_file"], data_only=True)
        target_wb = openpyxl.load_workbook(config["target_file"])

        source_ws = source_wb[config["source_sheet"]]
        source_ws_highlight = source_wb_highlight[config["source_sheet"]]
        target_ws = target_wb[config["target_sheet"]]

        source_header_row = find_first_header_row(source_ws, "Period")
        target_header_row = find_first_header_row(target_ws, "Period")

        source_headers = get_headers(source_ws, source_header_row)
        target_headers = get_headers(target_ws, target_header_row)

        # end_date = last recorded month in source
        source_period_col = source_headers.get(normalize("Period"))

        if not source_period_col:
            raise ValueError("Could not find Period column in source file")

        source_months = []

        for source_row in range(source_header_row + 1, source_ws.max_row + 1):
            period_value = source_ws.cell(row=source_row, column=source_period_col).value
            period_date = parse_period(period_value)

            if period_date is not None:
                source_months.append(first_day_of_month(period_date))

        if not source_months:
            raise ValueError("No valid periods found in source file")

        end_date = max(source_months)
        start_date = subtract_months(end_date, config["lookback_months"])

        new_columns = add_new_source_columns_to_output(
            source_ws,
            target_ws,
            source_header_row,
            target_header_row,
            source_headers,
            target_headers,
            start_date,
            end_date
        )

        for source_row in range(source_header_row + 1, source_ws.max_row + 1):  # locate header row and read onwards
            period_value = source_ws.cell(row=source_row, column=source_period_col).value
            period_date = parse_period(period_value)

            if period_date is None:
                continue

            period_month = first_day_of_month(period_date)

            if not (start_date <= period_month <= end_date):
                continue

            target_row = find_period_row(target_ws, target_headers, period_value)

            set_value(
                target_ws,
                target_row,
                target_headers,
                "Period",
                period_value,
                skipped_columns
            )

            for target_col, source_col in config["column_map"].items():
                source_value = get_value(
                    source_ws,
                    source_row,
                    source_headers,
                    source_col,
                    skipped_columns
                )

                target_col_num = target_headers.get(normalize(target_col))

                if not target_col_num:
                    skipped_columns.add(target_col)
                    continue

                existing_target_value = resolve_cell_value(
                    target_ws,
                    target_ws.cell(row=target_row, column=target_col_num).value
                )

                if (
                    target_col not in new_columns
                    and existing_target_value not in ("", None)
                    and source_value not in ("", None)
                    and values_different(existing_target_value, source_value)
                ):
                    diff = numeric_difference(existing_target_value, source_value)

                    # automatically update if the numeric difference is less than 1
                    if diff is None or diff >= 1:
                        flag_inconsistency(
                        target_ws,
                        source_ws_highlight,
                        target_row,
                        target_col_num,
                        source_row,
                        source_headers.get(normalize(source_col)),
                        inconsistencies,
                        period_value,
                        target_col,
                        existing_target_value,
                        source_value
                    )
                        continue
                
                # make sure it doesn't override existing 0 with blank if new building
                if source_value in ("", None):
                    pass
                else:
                    target_ws.cell(row=target_row, column=target_col_num).value = source_value

        save_changed_sheet_only(
            target_wb,
            config["target_sheet"],
            config["target_output_file"]
        )
        
        source_wb_highlight.save(config["source_output_file"])

        print(f"\nSuccess. Saved as {config['target_output_file']}")
        print(f"Highlighted source copy saved as {config['source_output_file']}")

        print(
            f"This application compared and transferred data only within the "
            f"last {config['lookback_months']} months "
            f"({start_date.strftime('%b %y')} to {end_date.strftime('%b %y')})."
        )

        if new_columns:
            print("\nNew source columns detected and added to output:")
            for col in new_columns:
                print(f"- {col}")
                print(
                "Previous months were filled with 0, and available source data "
                "has already been updated in the output Excel file."
                )

        if skipped_columns:
            print("\nMissing columns (possibly due to naming error):")
            for col in sorted(skipped_columns):
                print(f"- {col}")
        else:
            print("\nNo columns were skipped.")

        if inconsistencies:
            print("\n=============")
            print("\nNOTE:")
            print("Cells with inconsistencies were NOT overwritten.")
            print("The existing values in the target workbook were preserved.")

            print("\nInconsistencies found:")

            print("\nSUST office reference - output workbook cell numbers:")
            for item in inconsistencies:
                print(
                    f"- {item['period']} | {item['column']} ({item['target_cell']}): "
                    f"SUST had {item['existing_target_value']}, "
                    f"source has {item['new_source_value']}"
                )

            print("\nOther department reference - source workbook cell numbers:")
            for item in inconsistencies:
                print(
                    f"- {item['period']} | {item['column']} ({item['source_cell']}): "
                    f"SUST had {item['existing_target_value']}, "
                    f"source has {item['new_source_value']}"
                )

        else:
            print("\nNo inconsistencies found.")

    # for recyclable waste (iterate through sheets as months)

    def transfer_recyclable_wastes(config):
        skipped_columns = []

        source_wb = openpyxl.load_workbook(config["source_file"], data_only=True)
        target_wb = openpyxl.load_workbook(config["target_file"])

        source_ws = source_wb[config["source_sheet"]]
        target_ws = target_wb[config["target_sheet"]]

        target_header_row = 1
        target_headers = get_headers(target_ws, target_header_row)

        period_value = config["source_sheet"]
        target_row = find_month_row(target_ws, period_value)

        set_value(
            target_ws,
            target_row,
            target_headers,
            "Period",
            period_value,
            set()
        )

        for target_col, source_categories in config["column_map"].items():

            if isinstance(source_categories, list):
                source_value = sum(
                convert_to_kg(
                    source_category,
                    extract_total_by_category(
                        source_ws,
                        source_category,
                        config["source_total_header"]
                    )
                )
                for source_category in source_categories
            )

            else:
                source_value = extract_total_by_category(
                    source_ws,
                    source_categories,
                    config["source_total_header"]
                )
            
            source_value = convert_to_kg(source_categories, source_value)

            target_col_num = target_headers.get(normalize(target_col))

            if not target_col_num:
                skipped_columns.append(target_col)
                continue

            target_ws.cell(row=target_row, column=target_col_num).value = source_value

        save_changed_sheet_only(
            target_wb,
            config["target_sheet"],
            config["target_output_file"]
        )

        print(f"\nSuccess. Saved as {config['target_output_file']}")

        if skipped_columns:
            print("\nMissing target columns:")
            for col in skipped_columns:
                print(f"- {col}")
    
    category = config["category"].lower()

    if category in ("recyclable_wastes"):
        transfer_recyclable_wastes(config)
    else:
        transfer_recent_months(config)