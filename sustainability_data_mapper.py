import openpyxl
import re
from datetime import datetime
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
import operator
from openpyxl.styles import PatternFill

SOURCE_FILE = "testing_electricity_202604.xlsx"
TARGET_FILE = "testing_sustainability_data_report.xlsx"
TARGET_OUTPUT_FILE = "testing_sustainability_data_report_updated.xlsx"
SOURCE_OUTPUT_FILE = "testing_electricity_202604_highlighted.xlsx"

SOURCE_SHEET = "Breakdown"
TARGET_SHEET = "Monthly Electricity (Overview)"

DISCREPANCY_FILL = PatternFill(
    fill_type = "solid",
    start_color="FFFF00",
    end_color="FFFF00"
)

# system calculates 
OPERATORS = {
    "+": operator.add,
    "-": operator.sub,
    "*": operator.mul,
    "/": operator.truediv,
}


def resolve_cell_value(ws, value):
    if isinstance(value, str) and value.startswith("="):
        return calculate_simple_formula(ws, value)

    return value


def calculate_simple_formula(ws, formula):
    original_formula = formula

    formula = formula.strip().lstrip("=")

    # handles number related formulas in excel
    if re.fullmatch(r"\s*-?\d+(?:\.\d+)?(?:\s*[+\-*/]\s*-?\d+(?:\.\d+)?)+\s*", formula):
        try:
            return eval(formula, {"__builtins__": {}})
        except Exception:
            return original_formula

    # handles position related formulas in excel e.g. =Y146-Z146
    match = re.fullmatch(r"\s*([A-Z]+\d+)\s*([+\-*/])\s*([A-Z]+\d+)\s*", formula)
    if match:
        left_cell, op, right_cell = match.groups()

        left_value = resolve_cell_value(ws, ws[left_cell].value)
        right_value = resolve_cell_value(ws, ws[right_cell].value)

        try:
            left_value = float(left_value)
            right_value = float(right_value)

            if op == "+":
                return left_value + right_value
            if op == "-":
                return left_value - right_value
            if op == "*":
                return left_value * right_value
            if op == "/":
                return left_value / right_value
        except Exception:
            return original_formula

    return original_formula

# lookback month: compare / transfer data within 12 months

LOOKBACK_MONTHS = 12

# strip data of everything so formatting doesn't affect

def normalize(text):
    if text is None:
        return ""
    text = str(text).lower().strip()
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text

# parse and match rows by period (just a safety measure in case the formatting is ever changed)

def parse_period(value):
    if value is None:
        return None

    if isinstance(value, datetime):
        return value

    text = str(value).strip().replace("Sept", "Sep")

    for fmt in ("%b-%y", "%b %y", "%Y-%m-%d", "%Y/%m/%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            pass

    return None


def normalize_period(value):
    parsed = parse_period(value)
    if parsed:
        return parsed.strftime("%b-%y").lower()
    return str(value).strip().lower()


def periods_match(a, b):
    return normalize_period(a) == normalize_period(b)


def first_day_of_month(date_value):
    return datetime(date_value.year, date_value.month, 1)


def subtract_months(date_value, months):
    year = date_value.year
    month = date_value.month

    for _ in range(months):
        if month == 1:
            year -= 1
            month = 12
        else:
            month -= 1

    return datetime(year, month, 1)


def find_first_header_row(ws, required_header="Period"):
    required = normalize(required_header)

    for row in ws.iter_rows():
        values = [normalize(cell.value) for cell in row]
        if required in values:
            return row[0].row

    raise ValueError(f"Could not find first header row containing {required_header}")


def get_headers(ws, header_row):
    headers = {}

    for cell in ws[header_row]:
        if cell.value:
            headers[normalize(cell.value)] = cell.column

    return headers


def get_value(ws, row_num, headers, header_name, skipped_columns):
    col = headers.get(normalize(header_name))

    if not col:
        skipped_columns.add(header_name)
        return ""

    value = ws.cell(row=row_num, column=col).value
    return "" if value is None else value


def set_value(ws, row_num, headers, header_name, value, skipped_columns):
    col = headers.get(normalize(header_name))

    if not col:
        skipped_columns.add(header_name)
        return

    ws.cell(row=row_num, column=col).value = value


def find_period_row(ws, headers, period_value):
    period_col = headers.get(normalize("Period"))

    if not period_col:
        raise ValueError("Could not find Period column in target file")

    for row in range(1, ws.max_row + 1):
        if periods_match(ws.cell(row=row, column=period_col).value, period_value):
            return row

    return ws.max_row + 1   # add new row if period not found


def values_different(a, b):
    if a in ("", None) and b in ("", None):
        return False

    try:
        return abs(float(a) - float(b)) > 0.01
    except (ValueError, TypeError):
        return str(a).strip() != str(b).strip()
    
def numeric_difference(a, b):
    try:
        return abs(float(a) - float(b))
    except (ValueError, TypeError):
        return None

# highlight inconsistency in target

def flag_inconsistency(
    target_ws,
    source_ws_highlight,
    target_row,
    target_col,
    source_row,
    source_col,
    inconsistencies,
    period_value,
    column_name,
    existing_value,
    new_value,
    second_source_col=None
):
    target_cell = target_ws.cell(row=target_row, column=target_col)
    source_cell = source_ws_highlight.cell(row=source_row, column=source_col)

    target_cell.fill = DISCREPANCY_FILL
    source_cell.fill = DISCREPANCY_FILL

    source_cells = [source_cell.coordinate]

    if second_source_col is not None:
        second_source_cell = source_ws_highlight.cell(row=source_row, column=second_source_col)
        second_source_cell.fill = DISCREPANCY_FILL
        source_cells.append(second_source_cell.coordinate)

    inconsistencies.append({
        "period": normalize_period(period_value),
        "column": column_name,
        "existing_target_value": existing_value,
        "new_source_value": new_value,
        "target_cell": target_cell.coordinate,
        "source_cell": source_cell.coordinate
    })

# new buildings may be added to Excel. automatically update new column to target
def add_new_source_columns_to_target(
    source_ws,
    target_ws,
    source_header_row,
    target_header_row,
    source_headers,
    target_headers
):
    new_columns = []

    mapped_source_columns = {normalize(source_col) for source_col in COLUMN_MAP.values()}
    mapped_source_columns.add(normalize("Period"))

    for cell in source_ws[source_header_row]:
        if not cell.value:
            continue

        source_header_name = str(cell.value).strip()
        normalized_source_name = normalize(source_header_name)

        if normalized_source_name in mapped_source_columns:
            continue

        if normalized_source_name not in target_headers:
            new_col_num = target_ws.max_column + 1
            target_ws.cell(row=target_header_row, column=new_col_num).value = source_header_name

            target_headers[normalized_source_name] = new_col_num
            COLUMN_MAP[source_header_name] = source_header_name
            new_columns.append(source_header_name)

            for row in range(target_header_row + 1, target_ws.max_row + 1):
                if target_ws.cell(row=row, column=new_col_num).value in ("", None):
                    target_ws.cell(row=row, column=new_col_num).value = 0

    return new_columns

# copy columns based on matching

COLUMN_MAP = {
    "CLP Total": "CLP Total",
    "Main Academic Buildings & Other Public Facilities": "Main Academic Buildings & Other Public Facilities",
    "EC": "EC (ALL)",
    "CYT": "CYT",
    "Shaw Auditorium": "Shaw Auditorium",
    "APCF (R/F Zone L)": "APCF (R/F Zone L)",
    "New Data Centre (2/F Zone L)": "New Data Centre (2/F Zone L)",
    "Main Chiller Plant": "Main Chiller Plant",
    "LSKBB": "LSKBB",
    "IAS": "IAS",
    "Conference Lodge": "Conference Lodge",
    "LSK Chiller Plant": "LSK Chiller Plant",
    "WWT": "WWT",
    "Water Sport Centre": "Water Sport Centre",
    "SISC": "SISC",
    "HPC5": "HPC5",
    "RB2": "RB2",
    "IB": "IB",
    "SRD (UG10-13)": "SRD (UG10-13)",
    "TA-B": "TA-B",
    "GGT": "GGT",
    "Student Housing UG1-UG9, PG1-PG2": "Student Housing UG1-UG9, PG1-PG2",
    "Staff Quarters & Commercial Outlets": "Staff Quarters & Commercial Outlets",
    "Commercial Outlets": "Commercial Outlets",
    "Staff Quarters": "Staff Quarters",
    "Combined Chiller Plants": "Combined Chiller Plants",
    "TKO JC Hall": "TKO JC Hall",
    "EC 3/F Data Centre (Elect)": "EC 3/F Data Centre (Elect)",
    "EC 3/F Data Centre (Chiller kWhe)": "EC 3/F Data Centre (Chiller kWhe)",
    "SRD (Chiller kWhe)": "SRD (Chiller kWhe)",
    "IB (Chiller kWhe)": "IB (Chiller kWhe)",
    "BMW EV Charger kWh (LG7 & WWT)": "BMW EV Charger kWh (LG7 & WWT)",
    "Corner Stone EV Charger kWh": "Corner Stone EV Charger kWh"
}

def to_number(value):
    if value in ("", None):
        return 0

    try:
        return float(value)
    except (ValueError, TypeError):
        return 0

def transfer_recent_months(lookback_months):
    skipped_columns = set()
    inconsistencies = []

    today = datetime.today()
    end_date = first_day_of_month(today)
    start_date = subtract_months(end_date, lookback_months)

    source_wb = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    source_wb_highlight = openpyxl.load_workbook(SOURCE_FILE, data_only=True)
    target_wb = openpyxl.load_workbook(TARGET_FILE)

    source_ws = source_wb[SOURCE_SHEET]
    source_ws_highlight = source_wb_highlight[SOURCE_SHEET]
    target_ws = target_wb[TARGET_SHEET]

    source_header_row = find_first_header_row(source_ws, "Period")
    target_header_row = find_first_header_row(target_ws, "Period")

    source_headers = get_headers(source_ws, source_header_row)
    target_headers = get_headers(target_ws, target_header_row)

    new_columns = add_new_source_columns_to_target(
        source_ws,
        target_ws,
        source_header_row,
        target_header_row,
        source_headers,
        target_headers
    )

    source_period_col = source_headers.get(normalize("Period"))

    if not source_period_col:
        raise ValueError("Could not find Period column in source file")

    for source_row in range(source_header_row + 1, source_ws.max_row + 1):  # locate header row and read onwards
        period_value = source_ws.cell(row=source_row, column=source_period_col).value
        period_date = parse_period(period_value)

        if period_date is None:
            continue

        period_month = first_day_of_month(period_date)

        if not (start_date <= period_month <= end_date):
            continue

        target_row = find_period_row(target_ws, target_headers, period_value)

        set_value(
            target_ws,
            target_row,
            target_headers,
            "Period",
            period_value,
            skipped_columns
        )

        for target_col, source_col in COLUMN_MAP.items():
            source_value = get_value(
                source_ws,
                source_row,
                source_headers,
                source_col,
                skipped_columns
            )

            target_col_num = target_headers.get(normalize(target_col))

            if not target_col_num:
                skipped_columns.add(target_col)
                continue

            existing_target_value = resolve_cell_value(
                target_ws,
                target_ws.cell(row=target_row, column=target_col_num).value
            )

            if (
                target_col not in new_columns
                and existing_target_value not in ("", None)
                and source_value not in ("", None)
                and values_different(existing_target_value, source_value)
            ):
                diff = numeric_difference(existing_target_value, source_value)

                # automatically update if the numeric difference is less than 1
                if diff is None or diff >= 1:
                    flag_inconsistency(
                    target_ws,
                    source_ws_highlight,
                    target_row,
                    target_col_num,
                    source_row,
                    source_headers.get(normalize(source_col)),
                    inconsistencies,
                    period_value,
                    target_col,
                    existing_target_value,
                    source_value
                )
                    continue

            target_ws.cell(row=target_row, column=target_col_num).value = source_value

    target_wb.save(TARGET_OUTPUT_FILE)
    source_wb_highlight.save(SOURCE_OUTPUT_FILE)

    print(f"\nSuccess. Saved as {TARGET_OUTPUT_FILE}")
    print(f"Highlighted source copy saved as {SOURCE_OUTPUT_FILE}")

    print("\nComparison scope:")
    print(
        f"This application compared and transferred data only within the "
        f"last {lookback_months} months "
        f"({start_date.strftime('%b %y')} to {end_date.strftime('%b %y')})."
    )

    if new_columns:
        print("\nNew source columns detected and added to target:")
        for col in new_columns:
            print(f"- {col}")
            print(
            "Previous months were filled with 0, and available source data "
            "has already been updated in the target workbook."
            )

    if skipped_columns:
        print("\nMissing columns (possibly due to naming error):")
        for col in sorted(skipped_columns):
            print(f"- {col}")
    else:
        print("\nNo columns were skipped.")

    if inconsistencies:
        print("\n=============")
        print("\nNOTE:")
        print("Cells with inconsistencies were NOT overwritten.")
        print("The existing values in the target workbook were preserved.")

        print("\nInconsistencies found:")

        print("\nSUST office reference - output workbook cell numbers:")
        for item in inconsistencies:
            print(
                f"- {item['period']} | {item['column']} ({item['target_cell']}): "
                f"SUST had {item['existing_target_value']}, "
                f"source has {item['new_source_value']}"
            )

        print("\nOther department reference - source workbook cell numbers:")
        for item in inconsistencies:
            print(
                f"- {item['period']} | {item['column']} ({item['source_cell']}): "
                f"SUST had {item['existing_target_value']}, "
                f"source has {item['new_source_value']}"
            )

    else:
        print("\nNo inconsistencies found.")


transfer_recent_months(LOOKBACK_MONTHS)